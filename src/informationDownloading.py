# playwright 
from playwright.sync_api import sync_playwright
from datetime import datetime
import pandas as pd
from pathlib import Path
import openpyxl
import os
import time
import locale
from usefulFunctions import *

class informationDownloading():

    def __init__(self):
        self.browser = None
        self.context = None
        self.page = None

        self.sgte = None
        self.table=[]
        self.xlsxNamesList = []
        
    
    def init_page(self):
        # Go to https://www.borrd.com/
        self.page.goto("http://sgv.grupo-venado.com/venado/login.jsf")
        self.page.get_by_placeholder("Usuario").click()
        self.page.get_by_placeholder("Usuario").fill("BOT.ADMINISTRACION.LP")
        self.page.get_by_placeholder("Contraseña").click()
        self.page.get_by_placeholder("Contraseña").fill("venadobot")
        self.page.get_by_role("button", name="Iniciar Sesión").click()
        self.page.wait_for_load_state()

    def goto_bills(self):
        self.page.get_by_role("link", name="  Cobranza").click()
        self.page.get_by_role("link", name=" Cierres de Caja").click()
        self.page.wait_for_load_state()

    def set_day(self, dExcel, cssDate):
        if cssDate=="input#startDate":
            dates=[x for x in self.page.query_selector_all("div:nth-child(10) div.datepicker-days tbody td[class='day']")]
        elif cssDate=="input#endDate":
            dates=[x for x in self.page.query_selector_all("div:nth-child(11) div.datepicker-days tbody td[class='day']")]
        else:
            return
        for d in dates:
            if d.inner_text()==dExcel.strftime("%d"):
                d.click()
                break
    
    def tableCashClosing_and_download(self):
        
        headersTable=[x.inner_text() for x in self.page.query_selector_all("table#cashierClosings thead th")]
        rows=self.page.query_selector_all("table#cashierClosings tbody tr")
        print(len(rows))
        for row in rows:
            if len(row.query_selector_all("a"))==7:
                tipe="distribuidora"
            elif len(row.query_selector_all("a"))==5:
                tipe="agencia"
            else:
                tipe="otro"
            xpathArceoCajaBs="a[data-original-title='Arqueo de Caja Bs. EXCEL']"
            xpathArceoCajaUs="a[data-original-title='Arqueo de Caja $us. EXCEL']"
            xpathFirstExcel="a[data-original-title='Descargar EXCEL']"
            fields=[y.inner_text() for y in row.query_selector_all("td")]
            cashCode=fields[0]
            if tipe=="distribuidora":
                n1 = f"{cashCode}_arceoCajaBs.xls"
                n2 = f"{cashCode}_arceoCajaUs.xls"
                n3 = f"{cashCode}_firstExcel.xls"
                self.download_file(n1 ,xpathArceoCajaBs)
                self.download_file(n2 ,xpathArceoCajaUs)
                self.download_file(n3 ,xpathFirstExcel)
                self.xlsxNamesList.append(n1)
                self.xlsxNamesList.append(n2)
                self.xlsxNamesList.append(n3)

            elif tipe=="agencia":
                n4 = f"{cashCode}_arceoCajaBs.xls"
                n5 = f"{cashCode}_arceoCajaUs.xls"
                self.download_file(n4 ,xpathArceoCajaBs)
                self.download_file(n5 ,xpathArceoCajaUs)
                self.xlsxNamesList.append(n4)
                self.xlsxNamesList.append(n5)
            else:
                pass 
            
            rowDict={
                headersTable[0]:fields[0],
                headersTable[1]:fields[1],
                headersTable[2]:fields[2],
                headersTable[3]:fields[3],
                headersTable[4]:fields[4],
                headersTable[5]:fields[5],
                headersTable[6]:fields[6],
                headersTable[7]:fields[7],
                headersTable[8]:fields[8],
                headersTable[9]:tipe
            }
            self.table.append(rowDict)
        return pd.DataFrame(self.table)
    
    def evaluate_month(self, monthdate_obj, dExcel, cssDate):
        tday=dExcel.strftime("%B %Y")
        if monthdate_obj.strftime("%B %Y")==tday:
            print("same month")
            self.set_day(dExcel,cssDate)
            return True
        elif monthdate_obj<dExcel:
            print("next month")
            self.page.query_selector("div.datepicker-days th.next").click()
            return False
            #monthdate=w.find_element(By.CSS_SELECTOR,"div.datepicker-days th.datepicker-switch").text
        elif monthdate_obj>dExcel:
            print("previous month")
            self.page.query_selector("div.datepicker-days th.prev").click()
            return False

    def found_date(self, dExcel, cssDate):
        self.page.query_selector(cssDate).click()
        if cssDate=="input#startDate":
            monthdate=self.page.query_selector("body > div:nth-child(10) > div.datepicker-days > table > thead > tr:nth-child(1) > th.datepicker-switch").inner_text()
            monthdate=monthdate.replace("Septiembre","Setiembre")
            monthdate_obj=datetime.datetime.strptime(monthdate,"%B %Y")
        elif cssDate=="input#endDate":
            monthdate=self.page.query_selector("body > div:nth-child(11) > div.datepicker-days > table > thead > tr:nth-child(1) > th.datepicker-switch").inner_text()
            monthdate=monthdate.replace("Septiembre","Setiembre")
            print(monthdate)
            monthdate_obj=datetime.datetime.strptime(monthdate,"%B %Y")
    
        dateNotfound=True
        while dateNotfound:
            if self.evaluate_month(monthdate_obj,dExcel,cssDate):
                dateNotfound=False
                print("date found")
            else:
                monthdate=self.page.query_selector("div.datepicker-days th.datepicker-switch").inner_text()
                monthdate=monthdate.replace("Septiembre","Setiembre")
                monthdate_obj=datetime.datetime.strptime(monthdate,"%B %Y")
    
    def set_dates(self, dinit, dEnd):
        self.found_date(dinit,"input#startDate")
        time.sleep(1)
        self.found_date(dEnd,"input#endDate")
    
    def in_folder(self, nameFolder):
        folderParent = os.getcwd()
        #folderParent=Path(folderParent).parent
        folderParent=os.path.join(folderParent,nameFolder)
        return folderParent

    def download_file(self, nameFile, cssSelector):
        with self.page.expect_download() as download_info:
            self.page.query_selector(cssSelector).click()
        download = download_info.value
        nameFile=os.path.join(self.in_folder("descargas"),nameFile)
        download.save_as(nameFile)

    def download_AllFiles(self):
        maxPage = self.page.query_selector_all(".paginate_button")[7].inner_text()
        nClicks = int(maxPage) - 1
        self.tableCashClosing_and_download()
        for i in range(nClicks):
            self.sgte = "#cashierClosings_next a"
            self.page.query_selector(self.sgte).click()
            self.tableCashClosing_and_download()

    def main(self):
    
        bot2Path = currentPathParentFolder
        bot2Path = os.path.join(bot2Path, 'src')
        bot2Path = os.path.join(bot2Path, 'config2.xlsx')
        wb=openpyxl.load_workbook(bot2Path)
        ws=wb["Hoja1"]
        dinit=ws["B2"].value
        dEnd=ws["B3"].value
        locale.setlocale(locale.LC_TIME, '')
        with sync_playwright() as p:
            pw = informationDownloading()
            pw.browser = p.chromium.launch(headless=False)
            pw.context  = pw.browser.new_context ()
                    # Open new self.page
            pw.page = pw.context .new_page()
            pw.init_page()
            pw.goto_bills()
            pw.found_date(dinit, "input#startDate")
            time.sleep(1)
            pw.found_date(dEnd, "input#endDate")

            pw.download_AllFiles()
            # print(df)
            print(pd.DataFrame(pw.table))
            pw.page.pause()

if __name__ == "__main__":
    x = informationDownloading()
    x.main()